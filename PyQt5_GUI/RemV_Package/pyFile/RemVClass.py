# coding:utf-8
import os
import pickle
import re
import sys
import time
import webbrowser
import threading
from random import randint
from urllib import request

import openpyxl
import pymysql
import requests
from PyQt5 import QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon, QCursor
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QDialog, QWidget

import dataBase_Tools
import downloadScene
import getPronFromYouDao
import getSourceFromOuLu
import getSuffixFromCgdict
from pyFile import functions, getTranslationFromYouDao, createBookScene, addWordScene, \
    GUI_Chinese_Adjust, getGreatSentences

"""
@copyright   Copyright 2020 RemV
@license     GPL-3.0 (http://www.gnu.org/licenses/gpl-3.0.html)
@author      Lingao Xiao 肖凌奥 <920338028@qq.com>
@version     version 1.1
@link        https://github.com/ArmandXiao/RemV.git
"""


def toRelativePath(path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, path)


def loadQss():
    with open(toRelativePath("lib\\qss.txt"), "r") as f:
        return f.read()


def internetCheck():
    try:
        response = requests.get("http://www.baidu.com", timeout=2)
    except:
        return False
    return True


class RemVClass(QMainWindow):

    def __init__(self):
        super().__init__()

        self.dbStatus = False
        self.conn_user = None
        # self.conn_root = None
        self.cursor_user = None
        # self.cursor_root = None

        # 中文版
        self.ui = GUI_Chinese_Adjust.Ui_MainWindow()
        GUI_Chinese_Adjust.Ui_MainWindow.setupUi(self.ui, self)

        self.ui.stackedWidget.setCurrentIndex(4)

        # 禁止窗口缩放和拉伸
        self.setWindowFlag(Qt.FramelessWindowHint)

        # titleBar 图片
        self.ui.TitleBar_Widget.setStyleSheet("#TitleBar_Widget{border-image:url(./lib/res/image/background_5);}")
        self.ui.centralwidget.setStyleSheet("#centralwidget{border-image:url(./lib/res/image/background_yellow);}")

        self.setWindowIcon(QIcon("lib\\res\\image\\logo_1_128x128.ico"))
        self.ui.icoBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\logo_1_128x128.ico")))

        # 先disable
        self.ui.stackedWidget.setEnabled(False)

        # 给各个按钮加对应图标
        self.ui.uploadButton.setIcon(QIcon(toRelativePath("lib\\res\\image\\upload.png")))
        self.ui.MemorizeBtn_0.setIcon(QIcon(toRelativePath("lib\\res\\image\\brain.png")))
        self.ui.MenuBtn_1.setIcon(QIcon(toRelativePath("lib\\res\\image\\home_2.png")))
        self.ui.MenuBtn_2.setIcon(QIcon(toRelativePath("lib\\res\\image\\home_2.png")))
        self.ui.QuizBtn_0.setIcon(QIcon(toRelativePath("lib\\res\\image\\quiz.png")))
        self.ui.QuizBtn_1.setIcon(QIcon(toRelativePath("lib\\res\\image\\quiz.png")))
        self.ui.helpBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\question.png")))
        self.ui.translateBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\translate.png")))
        self.ui.translateBtn_2.setIcon(QIcon(toRelativePath("lib\\res\\image\\translate.png")))
        self.ui.backBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\back_2.png")))
        self.ui.NextBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\next_2.png")))
        self.ui.showBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\word.png")))
        self.ui.statusBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\wrong.png")))
        self.ui.exitBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\exit.png")))
        self.ui.miniBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\minimize.png")))
        self.ui.addBookBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\enter.png")))
        self.ui.proUSBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\US.png")))
        self.ui.proENGBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\ENG.png")))
        self.ui.proUSBtn_2.setIcon(QIcon(toRelativePath("lib\\res\\image\\US.png")))
        self.ui.proENGBtn_2.setIcon(QIcon(toRelativePath("lib\\res\\image\\ENG.png")))
        self.ui.starBtn_1.setIcon(QIcon(toRelativePath("lib\\res\\image\\star.png")))
        self.ui.starBtn_2.setIcon(QIcon(toRelativePath("lib\\res\\image\\star.png")))
        self.ui.starBtn_3.setIcon(QIcon(toRelativePath("lib\\res\\image\\star.png")))
        self.ui.starBtn_4.setIcon(QIcon(toRelativePath("lib\\res\\image\\star.png")))
        self.ui.starBtn_5.setIcon(QIcon(toRelativePath("lib\\res\\image\\star.png")))

        # hide widgets
        self.ui.starBtn_1.setVisible(False)
        self.ui.starBtn_2.setVisible(False)
        self.ui.starBtn_3.setVisible(False)
        self.ui.starBtn_4.setVisible(False)
        self.ui.starBtn_5.setVisible(False)

        self.ui.PreSufBox.setVisible(False)
        self.ui.PreSufBrowser.setVisible(False)

        self.ui.exchangeBox.hide()

        # 给列表添加 spacing
        self.ui.bookListWidget.setSpacing(20)
        self.ui.lessonListWidget.setSpacing(10)
        self.ui.meaningListWidget.setSpacing(5)
        self.ui.wordListWidget.setSpacing(5)

        # 给列表添加 点击事件
        self.ui.bookListWidget.itemClicked.connect(self.bookClicked)
        self.ui.bookListWidget.itemDoubleClicked.connect(self.bookDoubleClicked)
        self.ui.lessonListWidget.itemClicked.connect(self.lessonClicked)
        self.ui.wordListWidget.itemClicked.connect(self.relateMeaning)
        self.ui.meaningListWidget.itemClicked.connect(self.relateWord)
        # uploadBtn 添加点击事件
        self.ui.uploadButton.clicked.connect(self.uploadBtnClicked)
        self.ui.addBookBtn.clicked.connect(self.goToAddScene)
        self.ui.downloadBtn.clicked.connect(self.goToDownload)

        # 给 memorize, quiz, menu 添加事件 切屏事件
        self.ui.MenuBtn_1.clicked.connect(self.changeScene_0)
        self.ui.MenuBtn_2.clicked.connect(self.changeScene_0)
        self.ui.MemorizeBtn_0.clicked.connect(self.changeScene_1)
        self.ui.QuizBtn_0.clicked.connect(self.changeScene_2)
        self.ui.QuizBtn_1.clicked.connect(self.changeScene_2)
        self.ui.helpBtn.clicked.connect(self.help)

        # next, back, translate, show, exit, pronouce 添加点击事件
        self.ui.NextBtn.clicked.connect(self.next)
        self.ui.backBtn.clicked.connect(self.back)
        self.ui.translateBtn.clicked.connect(self.translateBtnClicked)
        self.ui.translateBtn_2.clicked.connect(self.translate)
        self.ui.showBtn.clicked.connect(lambda: self.updateWord(self.currentIndex))
        self.ui.exitBtn.clicked.connect(self.exitBtnClicked)
        self.ui.miniBtn.clicked.connect(lambda: self.showMinimized())
        self.ui.proUSBtn.clicked.connect(lambda: self.prounciate(self.currentWord, 0))
        self.ui.proENGBtn.clicked.connect(lambda: self.prounciate(self.currentWord, 1))
        self.ui.proUSBtn_2.clicked.connect(lambda: self.prounciate(self.currentWord, 0))
        self.ui.proENGBtn_2.clicked.connect(lambda: self.prounciate(self.currentWord, 1))

        self.ui.ChineseBtn.clicked.connect(self.ChineseBtnClicked)
        self.ui.EnglishBtn.clicked.connect(self.EnglishBtnClicked)

        # QuizScene 事件
        # 回车键叫return
        self.ui.enterEdit.returnPressed.connect(self.enterCheck)
        self.ui.enterEdit.textChanged.connect(self.checkEverySyllable)

        # MenuBar 事件
        self.ui.updateCheckBtn.pressed.connect(self.updateCheck)

        # add styleSheet for scrollBars
        self.setScrollBarStyle()

        self.pathList = ["lib\\res\\word_Repository\\SatVocabulary.xlsx"]

        # 总共有多少课
        self.lessonNum = 0
        # 装每节课的列表
        self.lessonList = []
        self.currentBook = ""
        self.currentLesson = 0
        self.currentSeq = "List Order"
        self.currentIndex = 0
        # words list for a specific book
        self.wordsLFSB = []  # words of all books
        self.wordsOAB = {}
        self.transList = {}
        # 当前数据
        self.currentWord = ""
        self.currentPOS = ""
        self.currentMeaning = ""
        self.countRound = 0
        self.lessonLen = 0
        # 本次学习数量
        self.nowNum = 0
        # 共学习数量
        self.accumulativeNum = 0
        # 控制是翻译还是显示原意思
        self.controlTranslate = True
        self.totalStudyTime = 0
        self.lastProgress = "欢迎使用这个软件，希望你可以喜欢"
        self.randomSet = set()
        self.noWrongTime = True
        self.remain = 0
        self.wrongSpel = False
        self.firstQuiz = True

        self.currentChineseTrans = ""
        self.currentEnglishTrans = ""

        # 0 is Chinese, 1 is English
        self.TransMode = 0

        self.transSourceControl = 1

        # Createbook
        self.createdBookName = ""
        # 获取上次进度
        try:
            self.getData()
        except:
            pass

        # Update the record of latest progress.
        self.ui.progressLabel.setText(self.lastProgress)

        # 解析已存在的excel
        '''
        Do not parse all excels at the beginning. It wastes memories and affects the speed.
        Parse the excel when user gives clear sign of opening a specific excel.
        '''
        # self.parseAllBooks(self.pathList)
        self.creatErrorBook()
        self.loadBookNames(self.pathList)

        if self.totalStudyTime == 0:
            QMessageBox.information(self, "介绍和使用说明", "致用户的一封信：\n\n\t欢迎使用RemV,"
                                                     "这是一款可以帮助你深度记忆单词的\n\t一款软件。"
                                                     "此软件通过与用户互动提高注意力,从而达\n\t到更好的记忆效果!\n\n"
                                                     "使用说明：\n\t1.上传文件或者使用本地提供的库。"
                                                     "\n\t2. 选择一个自动生成的Lesson。\n\t3. 点击\"背单词\"或\"小测\"按钮 \n\n\t"
                                                     "不再让英语成为负担, 祝你好运!\n\n肖凌奥 "
                                                     "Armand\n联系方式(微信): xla920338028")

        self.secondWin = self.accessSecond()()
        self.thirdWin = self.accessThird()()
        self.downloadScene = self.accessDownloadScene()()

        self.secondWin.setWindowTitle("创建书")
        self.thirdWin.setWindowTitle("加入单词")
        self.downloadScene.setWindowTitle("下载离线单词书")

        self.secondWin.setWindowIcon(QIcon("lib\\res\\image\\logo_1_128x128.ico"))
        self.thirdWin.setWindowIcon(QIcon("lib\\res\\image\\logo_1_128x128.ico"))
        self.downloadScene.setWindowIcon(QIcon("lib\\res\\image\\logo_1_128x128.ico"))

    def relateMeaning(self, item):
        index = self.ui.wordListWidget.row(item)
        self.ui.meaningListWidget.setCurrentRow(index)

    def relateWord(self, item):
        index = self.ui.meaningListWidget.row(item)
        self.ui.wordListWidget.setCurrentRow(index)

    def EnglishBtnClicked(self):
        self.TransMode = 1
        self.ui.EnglishBtn.setEnabled(False)
        self.ui.ChineseBtn.setEnabled(True)
        if self.currentEnglishTrans == "":
            self.ui.meaningBrowser.setText("No record found, sorry!")
        else:
            self.ui.meaningBrowser.setText(self.currentEnglishTrans)

    def ChineseBtnClicked(self):
        self.TransMode = 0
        self.ui.ChineseBtn.setEnabled(False)
        self.ui.EnglishBtn.setEnabled(True)
        # If no translation found, use original data
        if self.currentChineseTrans == "":
            self.ui.meaningBrowser.setText(self.currentMeaning)
            self.ui.meaningBrowser.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        else:
            self.ui.meaningBrowser.setText(self.currentChineseTrans)

    def exitBtnClicked(self):
        try:
            if self.conn_user:
                # self.conn_root.close()
                self.conn_user.close()

        finally:
            self.close()

    def help(self):
        responese = QMessageBox.question(self, "帮助", "联系方式: 920338028@qq.com (邮箱)"
                                                     "\n\t xla920338028 (微信)"
                                                     "\n "
                                                     "\n有任何使用问题请前往: "
                                                     "\n https://github.com/ArmandXiao/RemV.git"
                                                     "\n"
                                                     "\n - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - "
                                                     "\n - -                          是否前往？                        - -"
                                                     "\n - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - ")
        if responese != 65536:
            webbrowser.open("https://github.com/ArmandXiao/RemV.git")

    def loadBookNames(self, pathList):
        """
        Load book names but does not load the content of books.
        :param pathList:
        :return:
        """

        tmpList1 = []
        for path in pathList:
            tmpList1 += functions.getBookNames([path])

        # add book names in bookList
        self.ui.bookListWidget.addItems(tmpList1)

    def bookClicked(self, item):
        """
        bookListWidget的单机事件
        :param item: (点击自动提供) PyQt5.QtWidgets.QListWidgetItem object
        :return:
        """
        # self.ui.bookListWidget.row(item) 是获取所以索引

        self.ui.stackedWidget.setCurrentIndex(4)
        index = self.ui.bookListWidget.row(item)
        bookPath = (self.pathList[index])
        bool_ = True

        for each in self.wordsOAB.keys():
            if bookPath == each:
                bool_ = False
                break

        # If the book has been already parsed, do not parse it again.
        if (self.createdBookName == functions.getFileName(bookPath)) or (
                index == 0) or bool_:

            tmp = self.parseBook(bookPath)
            if tmp is not None:
                response = QMessageBox.question(self, "删除请求", "您是否需要从目录删除这本书")
                if response != 65536:
                    index = self.ui.bookListWidget.row(item)
                    self.pathList.pop(index)
                    self.ui.lessonListWidget.clear()
                    self.ui.bookListWidget.clear()
                    self.saveData()
                    self.getData()
                    self.loadBookNames(self.pathList)
                    self.ui.stackedWidget.setCurrentIndex(4)
                return

        if self.currentBook == bookPath:
            return
        # 更新 currentBook
        self.currentBook = bookPath
        # 更新这本书对应的课程
        self.setLessons(self.currentBook)

    def bookDoubleClicked(self, item):
        """
        Delete book from bookList
        :param item: the book item selected
        :return: None
        """

        reply = QMessageBox.question(self, "确认", "你确定要从目录删除这本书吗？",
                                     QMessageBox.Yes | QMessageBox.No)

        if reply != 65536:
            index = self.ui.bookListWidget.row(item)
            self.pathList.pop(index)
            self.ui.lessonListWidget.clear()
            self.ui.bookListWidget.clear()
            self.saveData()
            self.getData()
            self.loadBookNames(self.pathList)
            self.ui.stackedWidget.setCurrentIndex(4)
        else:
            return

    def lessonClicked(self, item):
        """
        lessonListWidget的单机事件
        :param item:(点击自动提供) PyQt5.QtWidgets.QListWidgetItem object
        :return:
        """
        # 判断有没有选择书
        if self.currentBook == "":
            return
        # check whether the item clicked is the first item, 'lessons' word.
        if self.ui.lessonListWidget.row(item) == 0:
            return
        else:
            self.ui.stackedWidget.setCurrentIndex(0)

        # 使右半边变成enabled
        self.ui.stackedWidget.setEnabled(True)
        # 更新 currentLesson, lessonLen
        # the reason to minus one is there is a 'lessons' word in position of index of 0 always.
        self.currentLesson = self.ui.lessonListWidget.row(item) - 1
        self.lessonLen = len(self.wordsOAB[self.currentBook][self.currentLesson])

        # 清空 wordListWi 和 meaningListWid
        self.ui.wordListWidget.clear()
        self.ui.meaningListWidget.clear()

        self.setOverViewScene()

    def setOverViewScene(self):
        """
        把OverViewScene设置好
        :return: None
        """

        self.ui.wordListWidget.clear()
        self.ui.meaningListWidget.clear()

        # 在测试完后 更新label
        self.ui.progressLabel.setText(self.lastProgress)

        # enable
        self.ui.translateBtn.setEnabled(True)
        self.ui.NextBtn.setEnabled(True)
        # 初始化第二轮的值
        self.currentIndex = 0
        self.countRound = 0

        self.ui.wordListWidget.setEnabled(True)
        self.ui.lessonListWidget.setEnabled(True)

        # 根据currentBook 和 currentLesson 填补单词和意思
        for i in range(self.lessonLen):
            if i == 0:  # update the first current word
                self.currentWord = self.wordsOAB[self.currentBook][self.currentLesson][i][0]

            self.ui.wordListWidget.addItem(
                self.wordsOAB[self.currentBook][self.currentLesson][i][0]
            )
            if self.transSourceControl:
                if i == 0:
                    self.currentMeaning = self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]
                self.ui.meaningListWidget.addItem(
                    self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]
                )
            elif (self.wordsOAB[self.currentBook][self.currentLesson][i][1][0] is not None) or (
                    self.wordsOAB[self.currentBook][self.currentLesson][i][1][0] != ""):
                if i == 0:  # update the first current meaning
                    self.currentMeaning = str(self.wordsOAB[self.currentBook][self.currentLesson][i][1][0]) + "  " + \
                                          self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]

                self.ui.meaningListWidget.addItem(
                    str(self.wordsOAB[self.currentBook][self.currentLesson][i][1][0]) + "  " +
                    self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]
                )
            else:
                if i == 0:
                    self.currentMeaning = self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]
                self.ui.meaningListWidget.addItem(
                    self.wordsOAB[self.currentBook][self.currentLesson][i][1][1]
                )

    def uploadBtnClicked(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "上传文件", "\\ ",
                                                  "Excel (*.xlsx);; CSV (*.csv)")  # 设置文件扩展名过滤,注意用双分号间隔
        # _ 是返回的type 如果是excel 就返回 "Excel (*.xlsx)"
        if _ != "":
            name = functions.getFileName(filePath)
            tmpList = functions.getBookNames(self.pathList)
            # 查重
            if name not in tmpList:
                self.pathList.append(filePath)
                self.loadBookNames([filePath])
            else:
                return
            self.saveData()
        else:
            print("上传动作取消")

    def changeScene_0(self):
        # disconnect database

        self.setOverViewScene()
        self.ui.stackedWidget.setCurrentIndex(0)
        self.ui.bookListWidget.setDisabled(False)
        self.ui.lessonListWidget.setDisabled(False)

    def changeScene_1(self):
        """
        Go to memorizing scene
        :return: None
        """
        self.connectDB()

        self.ui.stackedWidget.setCurrentIndex(1)
        self.ui.bookListWidget.setEnabled(False)
        self.ui.lessonListWidget.setEnabled(False)
        self.ui.QuizBtn_1.setVisible(False)
        self.ui.MenuBtn_1.setEnabled(True)

        # 把第一个元素更新
        self.updateWord(0)
        self.ui.countBrowser_1.setText("  1")
        self.ui.backBtn.setEnabled(False)
        # self.ui.showBtn.setVisible(False)

        self.currentIndex = 0

    def changeScene_2(self):
        """
        Go to test scene
        :return: None
        """

        self.connectDB()

        self.remain = self.lessonLen
        self.ui.stackedWidget.setCurrentIndex(2)
        self.ui.enterEdit.setEnabled(True)
        # reset firstQuiz to True
        self.firstQuiz = True

        self.nextRandWord()
        self.ui.enterEdit.setFocus()

        self.ui.statusBtn.setVisible(True)
        self.ui.bookListWidget.setEnabled(False)
        self.ui.lessonListWidget.setEnabled(False)
        self.ui.quizLabel.setText(
            "%s Lesson %d Quiz" % (functions.getBookNames([self.currentBook])[0], self.currentLesson + 1))

    def updateWord(self, index):
        """
        把 WordBrowser 和 meaningBrowser 更新
        :param index: 第几个数
        :return: None
        """
        # 更新 currentWord
        self.currentWord = self.wordsOAB[self.currentBook][self.currentLesson][index][0]

        self.ui.wordBrowser.setText(self.currentWord)

        if self.transSourceControl or self.wordsOAB[self.currentBook][self.currentLesson][index][1][0] is None:
            self.currentMeaning = self.wordsOAB[self.currentBook][self.currentLesson][index][1][1]
            self.ui.meaningBrowser.setText(
                self.wordsOAB[self.currentBook][self.currentLesson][index][1][1]
            )

        elif self.wordsOAB[self.currentBook][self.currentLesson][index][1][0] is not None:
            self.currentMeaning = str(self.wordsOAB[self.currentBook][self.currentLesson][index][1][0]) + \
                                  self.wordsOAB[self.currentBook][self.currentLesson][index][1][1]
            self.ui.meaningBrowser.setText(
                str(self.wordsOAB[self.currentBook][self.currentLesson][index][1][0]) +
                self.wordsOAB[self.currentBook][self.currentLesson][index][1][1]
            )

        # 居中显示
        self.ui.wordBrowser.setAlignment(Qt.AlignCenter)
        # self.ui.meaningBrowser.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

    def clearCurrentStatus(self):
        """
        Clear all temporary status for each word
        :return: None
        """

        self.ui.exchangeBox.hide()
        self.clearFrequency()
        self.clearTags()

        self.ui.PreSufBox.setVisible(False)
        self.ui.PreSufBrowser.setVisible(False)

        self.currentChineseTrans = ""
        self.currentEnglishTrans = ""

    def next(self):
        self.clearCurrentStatus()

        if len(self.wordsOAB) == 0:
            pass
        elif self.currentIndex < self.lessonLen - 1:
            self.ui.translateBtn.setEnabled(True)
            self.ui.showBtn.setEnabled(True)
            self.ui.MenuBtn_1.setVisible(False)
            self.ui.backBtn.setEnabled(True)

            # 第一轮要先 index+1 再update
            if self.countRound == 0:
                self.currentIndex += 1
                self.saveData()
                self.updateWord(self.currentIndex)
                # 更新count
                self.ui.countBrowser_1.setText("  " + str(self.currentIndex + 1))

            # 第二轮 要先update 再加一
            elif self.countRound == 1:
                self.currentIndex += 1
                if self.currentIndex == 0:
                    self.ui.backBtn.setEnabled(False)

                self.updateWord(self.currentIndex)
                # 不显示意思
                self.ui.meaningBrowser.setText("")
                # 更新count
                self.ui.countBrowser_1.setText("  " + str(self.currentIndex + 1))

        # 第一轮结束
        elif self.currentIndex == self.lessonLen - 1 and self.countRound == 0:
            self.ui.backBtn.setEnabled(False)

            self.currentIndex = -1  # 设置成-1 就可以先更新变量 再update了
            self.ui.wordBrowser.setText("确认环节")
            self.ui.meaningBrowser.setText("\t中文意思没有啦"
                                           "\n\t不过你同样可以点击显示按钮来查看"
                                           "\n\t准备好了吗？")
            self.ui.countBrowser_1.setText("")
            self.countRound = 1
            self.ui.showBtn.setEnabled(False)
            self.ui.translateBtn.setEnabled(False)

        # 第二轮结束
        elif self.currentIndex == self.lessonLen - 1 and self.countRound == 1:
            # self.currentIndex = 0

            self.ui.NextBtn.setEnabled(False)
            self.ui.wordBrowser.setText("\t快来小测吧~")
            self.ui.meaningBrowser.setText("\t检查单词拼写"
                                           "\n\t才是这个软件的核心")
            self.ui.backBtn.setEnabled(False)
            self.ui.translateBtn.setEnabled(False)
            self.ui.MenuBtn_1.setVisible(True)
            self.ui.QuizBtn_1.setVisible(True)

            self.conn_user.close()
            self.conn_user = None
            # self.conn_root.close()
            return

    def back(self):
        self.clearCurrentStatus()

        self.currentIndex -= 1

        if self.currentIndex == 0:
            self.ui.backBtn.setEnabled(False)
        # 更新count
        self.ui.countBrowser_1.setText("  " + str(self.currentIndex + 1))
        self.saveData()
        self.updateWord(self.currentIndex)

        if self.countRound == 1:
            self.ui.meaningBrowser.setText("")

    def nextRandWord(self):
        # 不是first time
        if self.noWrongTime and not self.firstQuiz:
            self.ui.wordEnterListWidget.addItem(self.currentWord)
            self.ui.wordEnterListWidget.scrollToBottom()

            self.randomSet.add(self.currentIndex)
            self.remain -= 1
            # 是firstTime
            self.nowNum += 1
            self.accumulativeNum += 1
            self.saveData()
        else:
            self.firstQuiz = False
            self.noWrongTime = True

        # reset randomSet is in function CheckSyllable
        if len(self.randomSet) == self.lessonLen:
            # 结束 Test scene
            self.ui.MenuBtn_2.setVisible(True)
            # 其实还需要判断 currentLesson 有没有越界
            self.currentIndex = 0
            self.countRound = 0
            # 更新界面
            self.ui.remainLabel.setText("Congratulations!")
            self.ui.statusBtn.setVisible(False)
            self.ui.enterEdit.setEnabled(False)
            self.ui.stackedWidget.setCurrentIndex(3)

            # 清空list里所有组件的
            self.ui.wordEnterListWidget.clear()
            self.ui.bookListWidget.setEnabled(True)
            self.ui.lessonListWidget.setEnabled(True)
            self.saveData()
            self.getData()

            return

        tmp = randint(0, self.lessonLen - 1)
        while tmp in self.randomSet:
            tmp = randint(0, self.lessonLen - 1)
        self.currentIndex = tmp
        self.currentWord = self.wordsOAB[self.currentBook][self.currentLesson][self.currentIndex][0]
        self.currentMeaning = self.wordsOAB[self.currentBook][self.currentLesson][self.currentIndex][1][1]
        # self.updateAll()
        self.updateAllTest()

    def updateAllTest(self):
        self.ui.meaningBrowser_2.setText(self.currentMeaning)
        self.ui.hintEdit.setText(self.convertTohint())
        self.ui.remainLabel.setText("Remain: %s" % str(self.remain))
        self.ui.meaningBrowser_2.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

    def convertTohint(self):
        tmp = self.currentWord[0:1]
        pattern = re.compile(r"[a-z][A-Z]*")
        tmp += pattern.sub("*", self.currentWord[1:])
        return tmp

    def enterCheck(self):
        """
        check the spelling of words.
        move to next word | stay this -> nextRand()
        :return: None
        """
        self.ui.bookListWidget.setDisabled(True)
        self.ui.lessonListWidget.setDisabled(True)
        self.ui.MenuBtn_2.setVisible(False)
        if self.ui.enterEdit.text().strip() == self.currentWord:
            # 对的话才让下一个词
            self.nextRandWord()
        else:
            # 错的话 就知道输入正确为止
            self.ui.hintEdit.setText(self.currentWord)
            self.noWrongTime = False
            self.wrongSpel = True

            # add the number in to ErrorBook in another thread
            t2 = threading.Thread(target=self.addToBook, args=(self.currentWord,))
            t2.setDaemon(True)
            t2.start()

        # clear Entry
        self.ui.enterEdit.clear()

    def checkEverySyllable(self):
        """
        Check the spelling whenever the content of the enterEdit changes.
        Clear the hint when it receives an input.
        :return:
        """

        # Check whether the spelling is wrong. If it is, show hint; change the word to
        if self.wrongSpel:
            self.ui.hintEdit.setText(self.currentWord)
            self.wrongSpel = False
        else:
            self.ui.hintEdit.setText(self.convertTohint())

        # Check spelling and update status icon
        if self.ui.enterEdit.text().strip() == self.currentWord:
            self.ui.statusBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\correct.png")))
        else:
            self.ui.statusBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\wrong.png")))

        # check the whether the last term
        if len(self.randomSet) == self.lessonLen:
            self.ui.hintEdit.clear()
            self.randomSet = set()
            self.ui.statusBtn.setIcon(QIcon(toRelativePath("lib\\res\\image\\correct.png")))

            engSentence, chiSentence = getGreatSentences.getSentences()
            self.ui.greatSentenceEng.setText(engSentence)
            self.ui.greatSentenceCHI.setText(chiSentence)

            self.conn_user.close()
            self.conn_user = None
            return

    def translateBtnClicked(self):
        """
        When Translate Button is clicked multiple tasks should run at the same time
        :return: None
        """
        self.translate()
        # MainThread
        # MainThread Functions must be executed before threads.
        self.setPreOrSuf()
        self.loadWordExchange()

        # start a thread of setFrequency
        t2 = threading.Thread(target=self.setFrequency)
        t2.setDaemon(True)
        t2.start()

        # start a thread of setTags
        t3 = threading.Thread(target=self.setTags)
        t3.setDaemon(True)
        t3.start()

    def translate(self):
        """
        get pronunciations, part of speech, and meanings from internet.
        :return:
        """

        def trans(self):
            ProunceList, MeaningList = getTranslationFromYouDao.translate(self.currentWord)

            tmp = "音标:"
            for each in ProunceList:
                tmp += each
            tmp += "\n"
            for each in MeaningList:
                tmp += each + "\n"
            self.ui.meaningBrowser.setText(tmp)
            self.ui.meaningBrowser_2.setText(tmp)

        def translateOnline(self):
            str_ = "音标:"
            if self.cursor_user is not None:
                sql = "SELECT phonetic FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()
                # data is stored in tuple
                if data:
                    str_ += "[%s]" % data[0]
                else:
                    str_ += "未找到音标"
                str_ += '\n'

                sql = "SELECT pos FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()
                if data and data[0] != "":
                    if "/" in data[0]:
                        tmpList = data[0].split("/")
                        for each in tmpList:
                            eachList = each.split(":")
                            if eachList[0] == "j":
                                str_ += "(" + "adj" + ".) : " + (str(eachList[1]) + "%") + " | "
                            else:
                                str_ += "(" + eachList[0] + ".) : " + (str(eachList[1]) + "%") + " | "

                    else:
                        eachList = data[0].split(":")
                        if eachList[0] == "j":
                            str_ += "(" + "adj" + ".) : " + (str(eachList[1]) + "%")
                        else:
                            str_ += "(" + eachList[0] + ".) : " + (str(eachList[1]) + "%")
                    str_ += '\n'

                sql = "SELECT translation FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()
                # data is stored in tuple
                if not data:  # if no data found in database then go search on the internet
                    trans(self)
                    return
                str_ += data[0]
                self.currentChineseTrans = str_

                sql = "SELECT definition FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()

                if data[0] != "":
                    self.currentEnglishTrans = data[0].replace("\n", "\n\n")

                if not self.TransMode:
                    self.ui.meaningBrowser.setText(self.currentChineseTrans)
                else:
                    if self.currentEnglishTrans == "":
                        self.ui.meaningBrowser.setText("No record found, sorry!")
                    else:
                        self.ui.meaningBrowser.setText(self.currentEnglishTrans)

                self.ui.meaningBrowser_2.setText(str_)

            else:
                trans(self)

        def translateFromCSV(self):
            str_ = "音标"
            # [word, pos, translation, phonetic, collins, tag, definition, exchange]
            pos = self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][0]
            translation = self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][1]
            phonetic = self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][2]
            definition = self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][5]

            # data is stored in tuple
            if phonetic and (phonetic != ""):
                str_ += "[%s]" % phonetic
            else:
                str_ += "未找到音标"
            str_ += '\n'

            if pos and pos != "":
                if "/" in pos:
                    tmpList = pos.split("/")
                    for each in tmpList:
                        eachList = each.split(":")
                        if eachList[0] == "j":
                            str_ += "(" + "adj" + ".) : " + (str(eachList[1]) + "%") + " | "
                        else:
                            str_ += "(" + eachList[0] + ".) : " + (str(eachList[1]) + "%") + " | "

                else:
                    eachList = pos.split(":")
                    if eachList[0] == "j":
                        str_ += "(" + "adj" + ".) : " + (str(eachList[1]) + "%")
                    else:
                        str_ += "(" + eachList[0] + ".) : " + (str(eachList[1]) + "%")
                str_ += '\n'

            self.currentChineseTrans = str_ + translation

            if definition and definition != "":
                self.currentEnglishTrans = definition.replace("\n", "\n\n")

            if not self.TransMode:
                self.ui.meaningBrowser.setText(self.currentChineseTrans)
            else:
                if self.currentEnglishTrans == "":
                    self.ui.meaningBrowser.setText("No record found, sorry!")
                else:
                    self.ui.meaningBrowser.setText(self.currentEnglishTrans)

            self.ui.meaningBrowser_2.setText(self.currentChineseTrans)

        # the self.transSourceControl whether to use which source, default is 0
        if not self.transSourceControl:
            # To start online connections only when csv file in local is not existent.
            if not internetCheck():
                self.ui.meaningBrowser.setText("请检查网络连接\n/(ㄒoㄒ)/~~")
                self.ui.meaningBrowser_2.setText("请检查网络连接\n/(ㄒoㄒ)/~~")
                return
            translateOnline(self)
        else:
            translateFromCSV(self)

        return

    def loadWordExchange(self):
        data = None
        if not self.transSourceControl:
            # To get frequency from database or online
            if not internetCheck():
                return
            if self.cursor_user is not None:
                sql = "SELECT exchange FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()
        else:
            # uniform the formate
            data = [self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][6]]

        str_ = ""
        if data:
            dataList = data[0].split("/")
            for each in dataList:
                eachList = each.split(":")
                if eachList[0] == "p":
                    str_ += "过去式: " + '\n' + eachList[1]
                elif eachList[0] == "d":
                    str_ += "过去分词: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "i":
                    str_ += "现在分词: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "3":
                    str_ += "三单: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "r":
                    str_ += "比较级: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "t":
                    str_ += "最高级: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "s":
                    str_ += "复数: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "0":
                    str_ += "原型: " + '\n' + eachList[1] + '\n'
                elif eachList[0] == "1":
                    str_ += "变形: " + '\n' + eachList[1] + '\n'
                str_ += '\n'

            if str_.strip() != "":
                self.ui.exchangeEdit.setText(str_)
                self.ui.exchangeBox.show()

    def parseBook(self, path):
        """
        parse the first sheet of a excel, and update bookList.
        :param path: the relative path of the book
        :return: None if this function runs properly, else return "Error"
        """

        # 处理words 把SAT单词书 分成好几节课 然后把SAT这真本书 放到wordsOAB 里面 名字与书的内容
        try:
            # check whether corresponding csv file exists
            if ((".xlsx" in path) and (
                    not os.path.exists("lib/res/word_Repository/csv/%s_remv.csv" % functions.getFileName(path)))) or (
                    path == self.pathList[0]):
                # after parsing books, books have to be divided into lessons
                self.wordsLFSB = functions.divideIntoLessons(functions.excelParse_xlrd(toRelativePath(path), 1))
                # 1 represents to download offline translation
                self.transSourceControl = 0
            elif ".csv" in path:
                self.wordsLFSB = functions.divideIntoLessons(functions.parseCsv(toRelativePath(path)))
                self.transList.update({path: functions.divideIntoLessons(
                    functions.parseCsv(path))})
                self.transSourceControl = 1
            else:
                # after parsing books, books have to be divided into lessons
                self.wordsLFSB = functions.divideIntoLessons(functions.excelParse_xlrd(toRelativePath(path), 0))
                # 0 represents to only parse the excel

                # Pattern: (word,(pos,translation,phonetic,collins,tag,definition,exchange))
                self.transList.update({path: functions.divideIntoLessons(
                    functions.parseCsv("lib/res/word_Repository/csv/%s_remv.csv"
                                       % functions.getFileName(path)))})

                self.transSourceControl = 1
        except:
            QMessageBox.about(self, "温馨提示", "目标路径中不存在该文件")
            return "Error"

        # 把每本书的链接 和 内容 用字典储存
        self.wordsOAB.update({path: self.wordsLFSB})

        # 保存数据
        self.saveData()

    def setPreOrSuf(self):
        """
        Set Prefix and Suffix for each word
        :return: None
        """
        if not internetCheck():
            return
        PreList, SufList = getSuffixFromCgdict.getPreOrSuf(self.currentWord)

        if len(PreList) != 0:
            self.ui.PreSufBox.setVisible(True)
            self.ui.PreSufBrowser.setVisible(True)
            self.ui.PreSufBrowser.clear()
            for eachPre in PreList:
                self.ui.PreSufBrowser.append(eachPre)
                self.ui.PreSufBrowser.append("\n")
            return

        if len(SufList) != 0:
            self.ui.PreSufBrowser.clear()
            self.ui.PreSufBox.setVisible(True)
            self.ui.PreSufBrowser.setVisible(True)
            for eachSuf in SufList:
                self.ui.PreSufBrowser.append(eachSuf)
                self.ui.PreSufBrowser.append("\n")
            return

    def setFrequency(self):
        """
        Set Frequency of use for each word
        :return: None
        """
        fre = -1
        if not self.transSourceControl:
            # To get frequency from database or online
            if not internetCheck():
                return
            if self.cursor_user is not None:
                sql = "SELECT collins FROM stardict WHERE word like '%s'" % self.currentWord
                self.cursor_user.execute(sql)
                data = self.cursor_user.fetchone()
                # data is stored in tuple
                if not data:
                    fre = getTranslationFromYouDao.getFrequency(self.currentWord)
                else:
                    fre = int(data[0])
            else:
                fre = getTranslationFromYouDao.getFrequency(self.currentWord)
        else:
            fre = int(self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][3])

        if fre == -1:  # getTranslationFromYouDao.getFrequency(self.currentWord) may gives -1
            return
        for i in range(fre):
            if i == 0:
                self.ui.starBtn_1.setVisible(True)
            if i == 1:
                self.ui.starBtn_2.setVisible(True)
            if i == 2:
                self.ui.starBtn_3.setVisible(True)
            if i == 3:
                self.ui.starBtn_4.setVisible(True)
            if i == 4:
                self.ui.starBtn_5.setVisible(True)

        return

    def clearFrequency(self):
        for i in range(5):
            if i == 0:
                self.ui.starBtn_1.setVisible(False)
            if i == 1:
                self.ui.starBtn_2.setVisible(False)
            if i == 2:
                self.ui.starBtn_3.setVisible(False)
            if i == 3:
                self.ui.starBtn_4.setVisible(False)
            if i == 4:
                self.ui.starBtn_5.setVisible(False)

        return

    def setTags(self):
        tagList = []
        if not self.transSourceControl:
            if not internetCheck():
                return
            # To start online connections only when csv file in local is not existent.
            tagList = getSourceFromOuLu.getTags(self.currentWord)
        else:
            tag = self.transList[self.currentBook][self.currentLesson][self.currentIndex][1][4]
            tagList = tag.split("/")
            index = 0
            for eachItem in tagList:
                for each in eachItem.split(" "):
                    if not (index < len(tagList)):
                        continue
                    if each == "toefl":
                        tagList[index] = "托福"
                    elif each == "ielts":
                        tagList[index] = "雅思"
                    elif each == "zk":
                        tagList[index] = '中考'
                    elif each == "gk":
                        tagList[index] = '高考'
                    elif each == "ky":
                        tagList[index] = '考研'
                    elif each == "cet4":
                        tagList[index] = '四级'
                    elif each == "cet6":
                        tagList[index] = '六级'
                    elif each == "gre":
                        tagList[index] = 'GRE'
                    elif each == "sat":
                        tagList[index] = 'SAT'
                    index += 1

        length = len(tagList)
        if length != 0:
            for i in range(length):
                if i == 0:
                    self.ui.tag_1.setText(tagList[i])
                elif i == 1:
                    self.ui.tag_2.setText(tagList[i])
                elif i == 2:
                    self.ui.tag_3.setText(tagList[i])
                elif i == 3:
                    self.ui.tag_4.setText(tagList[i])
        return

    def clearTags(self):
        self.ui.tag_1.setText("")
        self.ui.tag_2.setText("")
        self.ui.tag_3.setText("")
        self.ui.tag_4.setText("")

    def setLessons(self, path):
        """
        Divide books into lessons containing 20 words in each
        Every Lessons'name is a number, this avoids too many 'lesson' words appears when a book has more than 20 lessons.
        :param path
        :return: None
        """
        # initialize
        # lessonList needs to be initialized every time when book is clicked

        self.ui.lessonListWidget.clear()
        self.ui.lessonListWidget.addItem("Lessons")

        self.lessonNum = len(self.wordsOAB[path])

        if len(self.wordsOAB[path][0]) == 0:
            self.ui.lessonListWidget.clear()
            return

        self.lessonList = []
        for i in range(1, self.lessonNum + 1):
            self.lessonList.append(" L - " + str(i))
        self.ui.lessonListWidget.addItems(self.lessonList)

    # saveData in pickle
    def saveData(self):
        """
        first Datum：pathList
        second Datum：accumulativeNum
        third Datum：totalStudyTime
        fourth Datum：上次进度
        :return: None
        """

        try:
            with open(toRelativePath('myData.pickle'), 'wb') as handle:
                pickle.dump(self.pathList, handle, protocol=pickle.HIGHEST_PROTOCOL)
                pickle.dump(self.accumulativeNum, handle, protocol=pickle.HIGHEST_PROTOCOL)
                pickle.dump(self.totalStudyTime, handle, protocol=pickle.HIGHEST_PROTOCOL)
                # currentBook 是地址 currentLesson是下标 得加一
                self.lastProgress = "进度: %s Lesson %d  共学习: %d 个单词！" % (
                    functions.getBookNames([self.currentBook])[0], self.currentLesson + 1, self.accumulativeNum)
                pickle.dump(self.lastProgress, handle, protocol=pickle.HIGHEST_PROTOCOL)
        except:
            pass

    # getData in pickle
    def getData(self):
        try:
            # 不要变成绝对地址
            with open(toRelativePath('myData.pickle'), 'rb') as handle:
                self.pathList = pickle.load(handle)
                self.accumulativeNum = pickle.load(handle)
                self.totalStudyTime = pickle.load(handle)
                try:
                    self.lastProgress = pickle.load(handle)
                except:
                    pass
        except:
            pass

    def creatErrorBook(self):
        path = 'lib\\res\\word_Repository\\ErrorBook.xlsx'
        if path in self.pathList:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RemV_ErrorBook"

        wb.save(path)
        self.pathList.insert(0, path)

    def addToBook(self, word, index=0, name="ErrorBook"):
        """
        add a word in to a excel
        :param word: the word needed to be entered
        :param index: 0 is ErrorBook , len(self.pathList)-1 is newly added book
        :param name: the name of the excel
        :return: None
        """
        count = 0
        for i in functions.getBookNames(self.pathList):
            if self.createdBookName == i:
                index = count
            count += 1

        wb = openpyxl.load_workbook(self.pathList[index])
        ws = wb.active
        try:  # ws may be a blank page
            for eachRow in ws:
                if word == str(eachRow[0].value):
                    return
        except:
            pass

        data = [word, None]
        try:
            _, mean = getTranslationFromYouDao.translate(word)
            # deBug cell type error: a cell cannot take a two dimension array
            if mean:
                data.append(mean[0])
            else:
                data.append("NO MEANING FOUND")
        except:
            QMessageBox.about(self, "提示", "抱歉, 此功能必须需要网络连接。\n请检查您的网络连接")

        ws.append(data)
        wb.save("lib\\res\\word_Repository\\" + name + ".xlsx")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_drag = True

            self.m_DragPosition = event.globalPos() - self.pos()
            self.setCursor(QCursor(Qt.OpenHandCursor))

        if event.button() == Qt.RightButton:
            self.close()

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_drag:
            # 当左键移动窗体修改偏移值
            # QPoint
            # 实时计算窗口左上角坐标
            self.move(QMouseEvent.globalPos() - self.m_DragPosition)

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_drag = False
        self.setCursor(QCursor(Qt.ArrowCursor))

    def connectDB(self):
        self.connectDB_user()

        # self.connectDB_root()

    def connectDB_user(self):
        try:
            self.conn_user = pymysql.connect(
                host='192.168.1.101',
                port=3306,
                user="remv_user",
                passwd="iloveRemV",
                db='remv'
            )
            self.cursor_user = self.conn_user.cursor()
        except:
            pass

    # def connectDB_root(self):
    #     try:
    #         self.conn_root = pymysql.connect(
    #             host='192.168.1.101',
    #             port=3306,
    #             user="root",
    #             passwd=getPassword.getRootPassword(),
    #             db='remv'
    #         )
    #         self.cursor_root = self.conn_root.cursor()
    #     except:
    #         pass

    def updateCheck(self):
        currentVersion = ""
        newVersion = ""
        with open(r"lib/version.txt", "r") as f:
            currentVersion = f.readline().strip()
        try:
            res = request.urlopen(
                r"https://github.com/ArmandXiao/RemV/blob/master/PyQt5_GUI/RemV_Package/lib/version.txt",
                timeout=3)
            html = res.read().decode("utf-8")

            findVersion = re.compile(
                "(<td id=\"LC1\" class=\"blob-code blob-code-inner js-file-line\">)(.*?)(</td>)")
            newVersion = findVersion.findall(html)[0][1].strip()
            if newVersion != currentVersion:
                response = QMessageBox.question(self, "软件更新提示", "新版本: %s "
                                                                "\n当前版本:\n %s\n"
                                                                "\n"
                                                                "\n 是否前去官网查看或下载最新版"
                                                                "\n\n https://github.com/ArmandXiao/RemV.git\n"
                                                                "\n\n\t是否前往？"
                                                % (newVersion, currentVersion))
                if response != 65536:
                    webbrowser.open("https://github.com/ArmandXiao/RemV.git")
            else:
                response = QMessageBox.information(self, "软件更新提示", "当前版本: %s"
                                                                   "\n已经是最新般啦！"
                                                                   "\n请放心使用"
                                                   % currentVersion)
        except:
            response = QMessageBox.question(self, "提示", "自动检测版本更新失败: 连接超时"
                                                        "\n请前往官网进行查看:"
                                                        "\n\nhttps://github.com/ArmandXiao/RemV.git"
                                                        "\n\n\t是否前往？")
            if response != 65536:
                webbrowser.open("https://github.com/ArmandXiao/RemV.git")

    def prounciate(self, word, type_):
        t1 = threading.Thread(target=getPronFromYouDao.playSound, args=(word, type_))
        t1.setDaemon(True)
        t1.start()

    def goToAddScene(self):
        self.secondWin.show()
        self.secondWin.ui_CB.enterNameEdit.setFocus()

    def accessSecond(self):
        outterClass = self

        class DialogWin_1(QDialog):
            def __init__(self):
                super().__init__()
                self.ui_CB = createBookScene.Ui_Dialog()
                createBookScene.Ui_Dialog.setupUi(self.ui_CB, self)
                self.ui_CB.createBookBtn.clicked.connect(self.openBook)

            def openBook(self):
                name = self.ui_CB.enterNameEdit.text()
                self.creatBook(name)

            def creatBook(self, name):
                outterClass.createdBookName = name
                if name == "":
                    return

                path = "lib\\res\\word_Repository\\" + name + ".xlsx"
                if path in outterClass.pathList:
                    outterClass.secondWin.hide()
                    outterClass.thirdWin.show()
                    return

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = name

                wb.save(path)

                outterClass.pathList.append(path)

                outterClass.saveData()
                outterClass.secondWin.hide()
                outterClass.thirdWin.show()

        return DialogWin_1

    def accessThird(self):
        outterClass = self

        class DialogWin_2(QDialog):
            def __init__(self):
                super().__init__()
                self.ui_AW = addWordScene.Ui_Dialog()
                addWordScene.Ui_Dialog.setupUi(self.ui_AW, self)

                self.ui_AW.doneBtn.clicked.connect(self.closeWin)
                self.ui_AW.wordEnter.returnPressed.connect(self.enterPressed)

            def enterPressed(self):
                t1 = threading.Thread(target=outterClass.addToBook, args=(
                    self.ui_AW.wordEnter.text(), len(outterClass.pathList) - 1, outterClass.createdBookName))
                t1.setDaemon(True)
                t1.start()
                self.ui_AW.wordEnter.clear()

            def closeWin(self):
                self.enterPressed()
                outterClass.saveData()
                outterClass.getData()
                outterClass.ui.bookListWidget.clear()
                outterClass.loadBookNames(outterClass.pathList)
                outterClass.ui.stackedWidget.setCurrentIndex(4)
                outterClass.ui.lessonListWidget.clear()
                outterClass.thirdWin.hide()

        return DialogWin_2

    def goToDownload(self):
        self.downloadScene.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)

        self.downloadScene.ui_DS.stackedWidget.setCurrentIndex(0)
        self.downloadScene.list_ = []
        self.downloadScene.bookNum = 0
        self.downloadScene.bool_ = True
        self.downloadScene.time = 0
        self.downloadScene.myThread = None

        if not internetCheck():
            QMessageBox.information(self, "网络连接失败", "数据库请求连接失败\n请检查网络连接")
            return
        conn = pymysql.connect(host='192.168.1.101', port=3306, user='remv_user', passwd="iloveRemV", db='remv')
        cur = conn.cursor()
        for i in range(6):
            if i == 0:
                cur.execute("select book from bookcatalog where booktype like '小学'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_1.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_1.addItems(each)
            elif i == 1:
                cur.execute("select book from bookcatalog where booktype like '初中'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_2.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_2.addItems(each)
            elif i == 2:
                cur.execute("select book from bookcatalog where booktype like '高中'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_3.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_3.addItems(each)
            elif i == 3:
                cur.execute("select book from bookcatalog where booktype like '大学'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_4.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_4.addItems(each)
            elif i == 4:
                cur.execute("select book from bookcatalog where booktype like '留学'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_5.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_5.addItems(each)
            elif i == 5:
                cur.execute("select book from bookcatalog where booktype like '研究生'")
                data = cur.fetchall()
                if not data:
                    continue
                self.downloadScene.ui_DS.list_6.clear()
                for each in data:
                    self.downloadScene.ui_DS.list_6.addItems(each)
        conn.close()
        self.downloadScene.show()

    def accessDownloadScene(self):
        outterClass = self

        class DownloadScene(QWidget):

            # list_1 : 小学
            # list_2 : 初中
            # list_3 : 高中
            # list_4 : 大学
            # list_5 : 留学
            # list_6 : 研究生
            def __init__(self):
                super().__init__()
                self.list_ = []
                self.bookNum = 0
                self.bool_ = True
                self.time = 0
                self.myThread = None
                self.ui_DS = downloadScene.Ui_Form()
                downloadScene.Ui_Form.setupUi(self.ui_DS, self)

                self.ui_DS.downloadBtn.clicked.connect(self.goToConfirm)
                self.ui_DS.backBtn_DS.clicked.connect(lambda: self.ui_DS.stackedWidget.setCurrentIndex(0))
                self.ui_DS.confirmBtn.clicked.connect(self.download)

            def goToConfirm(self):
                self.list_ = []
                self.ui_DS.stackedWidget.setCurrentIndex(1)
                self.list_ += self.ui_DS.list_1.selectedItems()
                self.list_ += self.ui_DS.list_2.selectedItems()
                self.list_ += self.ui_DS.list_3.selectedItems()
                self.list_ += self.ui_DS.list_4.selectedItems()
                self.list_ += self.ui_DS.list_5.selectedItems()
                self.list_ += self.ui_DS.list_6.selectedItems()
                tmpList = []
                for each in self.list_:
                    if each.text().strip() != "":
                        tmpList.append(each.text())
                        self.bookNum += 1
                self.list_ = tmpList
                self.ui_DS.confirmList.addItems(tmpList)

                self.time = len(self.list_) * 30
                timeStr = "%d 分钟 %d 秒" % (self.time // 60, self.time % 60)

                self.ui_DS.confirmLabel.setText("共选择 %d 本书, 需要约 %s" % (len(self.list_, ), timeStr))
                self.ui_DS.confirmLabel.setAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            def download(self):
                # filter the csv that has been downloaded already
                for each in self.list_:
                    if os.path.exists(r'lib/res/word_Repository/csv/%s_remv.csv' % each):
                        self.list_.remove(each)
                        self.bookNum -= 1
                self.myThread = MyThread(self.list_, outterClass)
                self.myThread.setDaemon(True)
                self.myThread.start()

                def timeCheck(self):
                    self.ui_DS.stackedWidget.setCurrentIndex(2)
                    step = self.time / 100
                    progress = 0
                    while True:
                        if self.bookNum == 0:
                            return
                        time.sleep(step)
                        progress += step
                        self.ui_DS.progressBar.setValue(progress)
                        downloadProgress = len(self.myThread.finishedList) / self.bookNum * 100
                        if self.myThread is not None and downloadProgress > progress:
                            progress = downloadProgress
                            self.ui_DS.progressBar.setValue(downloadProgress)
                        if progress == 100:
                            return

                timeCheck(self)

                self.ui_DS.stackedWidget.setCurrentIndex(0)
                self.ui_DS.confirmList.clear()
                self.ui_DS.finishedList.clear()
                self.ui_DS.list_1.clearSelection()
                self.ui_DS.list_2.clearSelection()
                self.ui_DS.list_3.clearSelection()
                self.ui_DS.list_4.clearSelection()
                self.ui_DS.list_5.clearSelection()
                self.ui_DS.list_6.clearSelection()
                self.ui_DS.finishedList.addItems(self.myThread.finishedList)

                # reset book list
                outterClass.saveData()
                outterClass.getData()

                for each in self.myThread.finishedList:
                    outterClass.ui.bookListWidget.addItem(each)

        return DownloadScene

    def setScrollBarStyle(self):
        verticalScrollBarStyle = """
            QScrollBar{
                background: transparent;
                width: 12px;
                margin: 0px 0px 0px 0px;
                padding-top: 0px;
                padding-bottom: 0px;
                }   
                
            QScrollBar:handle {
                background: rgba(0, 0, 0, 50);
                width: 12px;
                border-radius: 6px;
                border: none;
                }
                
            QScrollBar::handle:hover {
                background: rgba(0, 0, 0, 100);
                }
                
            QScrollBar:sub-line {
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: top;
                }
            
            QScrollBar:add-line{
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: bottom;
                }
                              
        """

        horizontalScrollBarStyle = """
            QScrollBar:horizontal {
                border: none;
                background: none;
                height: 12px;
                margin: 0px 0px 0 0px;
            }
            
            QScrollBar::handle:horizontal {
                background: rgba(0, 0, 0, 50);
                width: 12px;
                border-radius: 6px;
                border: none;
            }
            
            QScrollBar::handle:hover {
                background: rgba(0, 0, 0, 100);
            }
            
            QScrollBar:sub-line {
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: left;
            }
            
            QScrollBar:add-line {
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: right;
            }


        """
        horizontalScrollBarStyle_bookList = """
            QScrollBar:horizontal {
                border: none;
                background: none;
                height: 15px;
                margin: 0px 0px 0 0px;
            }

            QScrollBar::handle:horizontal {
                background: rgba(41,51,60,150);
                width: 15px;
                border: none;
            }

            QScrollBar::handle:hover {
                background: rgba(0, 0, 0, 100);
            }

            QScrollBar:sub-line {
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: left;
            }

            QScrollBar:add-line {
                height: 12px;
                width: 10px;
                background: transparent;
                subcontrol-position: right;
            }


        """
        self.ui.bookListWidget.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.bookListWidget.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle_bookList)

        self.ui.lessonListWidget.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.lessonListWidget.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.wordListWidget.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.wordListWidget.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.meaningListWidget.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.meaningListWidget.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.wordEnterListWidget.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.wordEnterListWidget.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.meaningBrowser.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.meaningBrowser.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.meaningBrowser_2.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)
        self.ui.meaningBrowser_2.horizontalScrollBar().setStyleSheet(horizontalScrollBarStyle)

        self.ui.PreSufBrowser.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)

        self.ui.exchangeEdit.verticalScrollBar().setStyleSheet(verticalScrollBarStyle)


class MyThread(threading.Thread):
    def __init__(self, list_, outterClass):
        threading.Thread.__init__(self)
        self.list_ = list_
        self.outterClass = outterClass
        self.finishedList = []

    def run(self):
        for each in self.list_:
            name = dataBase_Tools.writeCSV_byName(each)
            self.outterClass.pathList.append(r'lib/res/word_Repository/csv/%s_remv.csv' % each)
            self.finishedList.append(name)

        return self.finishedList


"""
@copyright   Copyright 2020 RemV
@license     GPL-3.0 (http://www.gnu.org/licenses/gpl-3.0.html)
@author      Lingao Xiao 肖凌奥 <920338028@qq.com>
@version     version 1.1
@link        https://github.com/ArmandXiao/RemV.git
"""
